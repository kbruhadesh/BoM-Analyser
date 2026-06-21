"""
excel_exporter.py — Generates a 3-sheet Excel report in INR.
Uses openpyxl with conditional formatting.
"""
import io
import logging
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Colours
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
RED_FILL   = PatternFill("solid", fgColor="FFC7CE")
AMBER_FILL = PatternFill("solid", fgColor="FFEB9C")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
BOLD = Font(bold=True)
THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

INR_FORMAT = '₹#,##0.00'
INR_FORMAT_4 = '₹#,##0.0000'


def _header_row(ws, headers: list[str], row: int = 1) -> None:
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


def export_to_excel(
    results: list[dict],
    task_id: str,
    usd_inr_rate: float,
    audit_log: list[dict] | None = None,
) -> bytes:
    """
    Build a 3-sheet Excel workbook and return as bytes.

    results: list of OptimizationOutput-compatible dicts
    """
    wb = Workbook()

    # ── Sheet 1: Optimization Summary ─────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Optimization Summary"
    ws1.row_dimensions[1].height = 30

    headers1 = [
        "Line", "Component / MPN", "Qty", "Best Vendor",
        "Unit Price (₹)", "Total Cost (₹)", "Availability",
        "MOQ", "Savings vs Worst (₹)", "Recommendation"
    ]
    _header_row(ws1, headers1)

    total_cost = 0.0
    total_savings = 0.0

    for i, r in enumerate(results, start=2):
        ws1.cell(i, 1, i - 1)
        ws1.cell(i, 2, r.get("normalized_mpn") or r.get("component"))
        ws1.cell(i, 3, r.get("quantity_required"))
        ws1.cell(i, 4, r.get("best_vendor") or "N/A")

        up = r.get("best_unit_price_inr")
        tp = r.get("best_total_price_inr")
        sv = r.get("savings_vs_worst_inr", 0)

        up_cell = ws1.cell(i, 5, up)
        tp_cell = ws1.cell(i, 6, tp)
        sv_cell = ws1.cell(i, 9, sv)

        if up is not None:
            up_cell.number_format = INR_FORMAT_4
        if tp is not None:
            tp_cell.number_format = INR_FORMAT
            total_cost += tp
        if sv:
            sv_cell.number_format = INR_FORMAT
            total_savings += sv

        avail = r.get("availability", "")
        av_cell = ws1.cell(i, 7, avail)
        if avail == "In Stock":
            av_cell.fill = GREEN_FILL
        elif avail == "Out of Stock":
            av_cell.fill = RED_FILL
        else:
            av_cell.fill = AMBER_FILL

        ws1.cell(i, 8, r.get("moq"))
        ws1.cell(i, 10, r.get("recommendation_reason", ""))

        # Apply border
        for col in range(1, 11):
            ws1.cell(i, col).border = BORDER
            ws1.cell(i, col).alignment = Alignment(vertical="center")

    # Summary totals row
    summary_row = len(results) + 3
    ws1.cell(summary_row, 1, "TOTAL")
    ws1.cell(summary_row, 1).font = BOLD
    tc_cell = ws1.cell(summary_row, 6, total_cost)
    tc_cell.number_format = INR_FORMAT
    tc_cell.font = BOLD
    sv_total_cell = ws1.cell(summary_row, 9, total_savings)
    sv_total_cell.number_format = INR_FORMAT
    sv_total_cell.font = BOLD

    # Rate footnote
    ws1.cell(summary_row + 2, 1, f"USD/INR rate used: ₹{usd_inr_rate:.2f}  |  Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}")
    ws1.cell(summary_row + 2, 1).font = Font(italic=True, size=9, color="666666")

    _auto_width(ws1)

    # ── Sheet 2: Full Vendor Comparison ───────────────────────────────────────
    ws2 = wb.create_sheet("Vendor Comparison")
    headers2 = [
        "Component", "Vendor", "Part #",
        "Unit Price (₹)", "Total (₹)", "Stock Qty", "MOQ",
        "Availability", "Lead Time (wks)",
        "Price Break 5", "Price Break 10", "Price Break 100",
    ]
    _header_row(ws2, headers2)
    ws2.row_dimensions[1].height = 30

    row_idx = 2
    for r in results:
        comp = r.get("normalized_mpn") or r.get("component")
        qty = r.get("quantity_required", 1)
        best_vendor = r.get("best_vendor")
        all_vendors = r.get("all_vendors", [])

        for v in all_vendors:
            vname = v.get("vendor", "")
            is_best = vname == best_vendor

            ws2.cell(row_idx, 1, comp)
            ws2.cell(row_idx, 2, vname)
            ws2.cell(row_idx, 3, v.get("vendor_part_number", ""))

            up = v.get("unit_price_inr")
            up_cell = ws2.cell(row_idx, 4, up)
            if up is not None:
                up_cell.number_format = INR_FORMAT_4
                if is_best:
                    up_cell.fill = GREEN_FILL
                    up_cell.font = BOLD

            tp_cell = ws2.cell(row_idx, 5, round(up * qty, 2) if up else None)
            if up:
                tp_cell.number_format = INR_FORMAT

            ws2.cell(row_idx, 6, v.get("stock_qty", 0))
            ws2.cell(row_idx, 7, v.get("moq", 1))

            avail = v.get("availability", "")
            av_cell = ws2.cell(row_idx, 8, avail)
            if avail == "Out of Stock":
                av_cell.fill = RED_FILL

            ws2.cell(row_idx, 9, v.get("lead_time_weeks"))

            # Price breaks at qty 5, 10, 100
            breaks = v.get("price_breaks", [])
            break_map = {b.get("qty"): b.get("price_inr") for b in breaks if b.get("qty")}
            for col_offset, target_qty in enumerate([5, 10, 100]):
                val = break_map.get(target_qty)
                if val is None:
                    # Find nearest ladder ≤ target_qty
                    candidates = {q: p for q, p in break_map.items() if q <= target_qty}
                    val = break_map[max(candidates)] if candidates else None
                cell = ws2.cell(row_idx, 10 + col_offset, val)
                if val is not None:
                    cell.number_format = INR_FORMAT_4

            for col in range(1, 13):
                ws2.cell(row_idx, col).border = BORDER
                ws2.cell(row_idx, col).alignment = Alignment(vertical="center")

            row_idx += 1

    _auto_width(ws2)

    # ── Sheet 3: Audit Log ────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Audit Log")
    headers3 = ["Component", "Vendor", "Scrape Time", "Vendors Queried", "Cache Hit", "Status", "Error"]
    _header_row(ws3, headers3)

    if audit_log:
        for i, entry in enumerate(audit_log, start=2):
            ws3.cell(i, 1, entry.get("component", ""))
            ws3.cell(i, 2, entry.get("vendor", ""))
            ws3.cell(i, 3, entry.get("scraped_at", ""))
            ws3.cell(i, 4, entry.get("vendors_queried", ""))
            ws3.cell(i, 5, "Yes" if entry.get("cache_hit") else "No")
            ws3.cell(i, 6, entry.get("status", ""))
            ws3.cell(i, 7, entry.get("error", ""))
            for col in range(1, 8):
                ws3.cell(i, col).border = BORDER

    _auto_width(ws3)

    # Write to bytes buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
