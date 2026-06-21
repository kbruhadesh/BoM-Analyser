"""
test_backend.py — Unit + integration tests for BoM Analyzer backend.
Run with:  pytest tests/test_backend.py -v
"""
import sys
import os
import asyncio
import pytest

# Ensure backend/ is on sys.path
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))


# ══════════════════════════════════════════════════════════════════════════════
# 1.  BoM Parser
# ══════════════════════════════════════════════════════════════════════════════

class TestBomParser:

    def test_csv_standard(self):
        from parser.bom_parser import parse_bom
        csv = "Qty,MPN,Description,Manufacturer,Ref\n10,ATmega328P-PU,8-bit MCU,Microchip,U1\n5,NRF24L01+,RF module,Nordic,U2"
        items = parse_bom(csv, "csv")
        assert len(items) == 2
        assert items[0].quantity == 10
        assert items[0].manufacturer_part_number == "ATMEGA328P-PU"
        assert items[1].quantity == 5

    def test_csv_fuzzy_headers(self):
        from parser.bom_parser import parse_bom
        csv = "count,part_number,component,mfr,ref des\n20,STM32F103C8T6,ARM MCU,STMicro,U3"
        items = parse_bom(csv, "csv")
        assert len(items) == 1
        assert items[0].quantity == 20
        assert "STM32F103C8T6" in items[0].manufacturer_part_number

    def test_tsv_input(self):
        from parser.bom_parser import parse_bom
        tsv = "Qty\tMPN\tDescription\n50\tGRM188R61A106KE69D\t100uF Cap"
        items = parse_bom(tsv, "tsv")
        assert len(items) == 1
        assert items[0].quantity == 50

    def test_plain_text_with_qty_prefix(self):
        from parser.bom_parser import parse_bom
        txt = "10x ATmega328P\n5x NRF24L01+"
        items = parse_bom(txt, "text")
        assert len(items) == 2
        assert items[0].quantity == 10

    def test_plain_text_mpn_list(self):
        from parser.bom_parser import parse_bom
        txt = "ATmega328P\nSTM32F103C8T6\nNRF24L01+"
        items = parse_bom(txt, "text")
        assert len(items) == 3

    def test_auto_detect_csv(self):
        from parser.bom_parser import parse_bom, _detect_format
        csv = "Qty,MPN,Desc\n1,ABC123,test"
        assert _detect_format(csv) == "csv"

    def test_auto_detect_tsv(self):
        from parser.bom_parser import _detect_format
        tsv = "Qty\tMPN\tDesc\n1\tABC123\ttest"
        assert _detect_format(tsv) == "tsv"

    def test_empty_rows_skipped(self):
        from parser.bom_parser import parse_bom
        csv = "Qty,MPN,Description\n10,ATmega328P,MCU\n,,\n5,NRF24L01+,RF"
        items = parse_bom(csv, "csv")
        assert len(items) == 2

    def test_quantity_defaults_to_1_on_bad_value(self):
        from parser.bom_parser import _clean_qty
        assert _clean_qty("abc") == 1
        assert _clean_qty("") == 1
        assert _clean_qty("0") == 1     # 0 clamped to 1

    def test_quantity_parses_commas(self):
        from parser.bom_parser import _clean_qty
        assert _clean_qty("1,000") == 1000


# ══════════════════════════════════════════════════════════════════════════════
# 2.  MPN Normalizer
# ══════════════════════════════════════════════════════════════════════════════

class TestNormalizer:

    def test_strip_dip_suffix(self):
        from parser.normalizer import normalize_mpn
        r = normalize_mpn("ATmega328P-PU")
        assert r.normalized_mpn == "ATmega328P"
        assert r.confidence >= 0.90

    def test_strip_soic_suffix(self):
        from parser.normalizer import normalize_mpn
        r = normalize_mpn("LM358-SOIC")
        assert "SOIC" not in r.normalized_mpn

    def test_alias_resolution(self):
        from parser.normalizer import normalize_mpn
        r = normalize_mpn("328p")
        assert r.normalized_mpn == "ATmega328P"

    def test_alias_lm7805(self):
        from parser.normalizer import normalize_mpn
        r = normalize_mpn("LM7805")
        assert r.normalized_mpn == "L7805"

    def test_strip_manufacturer_prefix(self):
        from parser.normalizer import normalize_mpn
        r = normalize_mpn("Microchip ATmega328P-PU")
        assert "MICROCHIP" not in r.normalized_mpn.upper()

    def test_uppercase_output(self):
        from parser.normalizer import normalize_mpn
        r = normalize_mpn("atmega328p")
        # Should be uppercase or resolved alias
        assert r.normalized_mpn == r.normalized_mpn.upper() or r.normalized_mpn == "ATmega328P"

    def test_confidence_high_for_exact(self):
        from parser.normalizer import normalize_mpn
        r = normalize_mpn("STM32F103C8T6")
        assert r.confidence >= 0.95

    def test_confidence_lower_after_alias(self):
        from parser.normalizer import normalize_mpn
        r = normalize_mpn("mega328")
        assert r.confidence < 1.0

    def test_empty_mpn_returns_zero_confidence(self):
        from parser.normalizer import normalize_mpn
        r = normalize_mpn("")
        assert r.confidence == 0.0

    def test_add_custom_alias(self):
        from parser.normalizer import normalize_mpn, add_alias
        add_alias("mytestpart123", "CANONICAL-MPN-XYZ")
        r = normalize_mpn("mytestpart123")
        assert r.normalized_mpn == "CANONICAL-MPN-XYZ"

    def test_indian_market_alias_ch340(self):
        from parser.normalizer import normalize_mpn
        r = normalize_mpn("CH340")
        assert r.normalized_mpn == "CH340G"

    def test_indian_market_alias_l298n(self):
        from parser.normalizer import normalize_mpn
        r = normalize_mpn("L298N")
        assert r.normalized_mpn == "L298N"


# ══════════════════════════════════════════════════════════════════════════════
# 3.  Cost Optimizer
# ══════════════════════════════════════════════════════════════════════════════

SAMPLE_VENDORS = [
    {"vendor": "LCSC",   "vendor_part_number": "C1570968", "unit_price_inr": 81.57, "stock_qty": 12000, "moq": 10,  "availability": "In Stock",     "lead_time_weeks": None, "price_breaks": []},
    {"vendor": "DigiKey","vendor_part_number": "ATMEGA328P-PU-ND", "unit_price_inr": 104.37, "stock_qty": 5000,  "moq": 1,   "availability": "In Stock",     "lead_time_weeks": None, "price_breaks": []},
    {"vendor": "Mouser", "vendor_part_number": "556-ATmega328P-PU","unit_price_inr": 108.55, "stock_qty": 3200,  "moq": 1,   "availability": "In Stock",     "lead_time_weeks": None, "price_breaks": []},
    {"vendor": "Arrow",  "vendor_part_number": "ATmega328P-PU",    "unit_price_inr": 112.62, "stock_qty": 0,     "moq": 1,   "availability": "Out of Stock", "lead_time_weeks": 8,    "price_breaks": []},
]


class TestCostOptimizer:

    def test_best_vendor_selected(self):
        from optimizer.cost_optimizer import optimize
        result = optimize("ATmega328P", "ATmega328P", SAMPLE_VENDORS, qty=10)
        # LCSC has lowest price but MOQ=10 which equals qty — should still win or DigiKey
        assert result.best_vendor in ("LCSC", "DigiKey")

    def test_oos_vendor_not_selected_as_best_when_stock_available(self):
        from optimizer.cost_optimizer import optimize
        result = optimize("ATmega328P", "ATmega328P", SAMPLE_VENDORS, qty=10)
        assert result.best_vendor != "Arrow"

    def test_savings_calculated(self):
        from optimizer.cost_optimizer import optimize
        result = optimize("ATmega328P", "ATmega328P", SAMPLE_VENDORS, qty=10)
        assert result.savings_vs_worst_inr >= 0

    def test_savings_equals_price_diff_times_qty(self):
        from optimizer.cost_optimizer import optimize, score_vendors
        result = optimize("ATmega328P", "ATmega328P", SAMPLE_VENDORS, qty=10)
        scored = score_vendors(SAMPLE_VENDORS, 10)
        in_stock_prices = [s.unit_price_inr for s in scored if s.availability == "In Stock"]
        expected_savings = round((max(in_stock_prices) - min(in_stock_prices)) * 10, 2)
        assert abs(result.savings_vs_worst_inr - expected_savings) < 0.10

    def test_no_vendors_returns_no_data(self):
        from optimizer.cost_optimizer import optimize
        result = optimize("TestPart", "TestPart", [], qty=5)
        assert result.best_vendor is None
        assert result.availability == "No Data"

    def test_all_oos_returns_oos_status(self):
        from optimizer.cost_optimizer import optimize
        all_oos = [dict(v, availability="Out of Stock", stock_qty=0) for v in SAMPLE_VENDORS]
        result = optimize("TestPart", "TestPart", all_oos, qty=5)
        assert result.availability == "Out of Stock"

    def test_split_order_when_no_single_vendor_has_full_stock(self):
        from optimizer.cost_optimizer import optimize
        low_stock = [
            {"vendor": "LCSC",   "vendor_part_number": "X1", "unit_price_inr": 80.0, "stock_qty": 5,  "moq": 1, "availability": "In Stock", "lead_time_weeks": None, "price_breaks": []},
            {"vendor": "DigiKey","vendor_part_number": "X2", "unit_price_inr": 95.0, "stock_qty": 5,  "moq": 1, "availability": "In Stock", "lead_time_weeks": None, "price_breaks": []},
        ]
        result = optimize("TestPart", "TestPart", low_stock, qty=9)
        # split_order should be populated
        assert result.split_order is not None
        assert len(result.split_order) >= 1

    def test_moq_penalty_applied(self):
        """Vendor with very high MOQ should score lower when qty is small."""
        from optimizer.cost_optimizer import score_vendors
        vendors = [
            {"vendor": "A", "vendor_part_number": "A1", "unit_price_inr": 50.0, "stock_qty": 1000, "moq": 500, "availability": "In Stock", "lead_time_weeks": None},
            {"vendor": "B", "vendor_part_number": "B1", "unit_price_inr": 55.0, "stock_qty": 1000, "moq": 1,   "availability": "In Stock", "lead_time_weeks": None},
        ]
        scored = score_vendors(vendors, qty=5)
        # B should rank higher because MOQ=1 fits qty=5 better, despite slightly higher price
        assert scored[0].vendor_name == "B"

    def test_recommendation_reason_populated(self):
        from optimizer.cost_optimizer import optimize
        result = optimize("ATmega328P", "ATmega328P", SAMPLE_VENDORS, qty=10)
        assert result.recommendation_reason
        assert len(result.recommendation_reason) > 5


# ══════════════════════════════════════════════════════════════════════════════
# 4.  Currency Converter
# ══════════════════════════════════════════════════════════════════════════════

class TestCurrency:

    def test_usd_to_inr_basic(self):
        from currency import usd_to_inr
        result = usd_to_inr(1.0, rate=83.5)
        assert abs(result - 83.5) < 0.01

    def test_inr_to_usd_basic(self):
        from currency import inr_to_usd
        result = inr_to_usd(83.5, rate=83.5)
        assert abs(result - 1.0) < 0.01

    def test_roundtrip(self):
        from currency import usd_to_inr, inr_to_usd
        rate = 83.5
        usd = 2.45
        inr = usd_to_inr(usd, rate)
        back = inr_to_usd(inr, rate)
        assert abs(back - usd) < 0.001

    def test_fallback_rate_is_positive(self):
        from currency import FALLBACK_USD_INR
        assert FALLBACK_USD_INR > 70.0   # sanity check for INR range


# ══════════════════════════════════════════════════════════════════════════════
# 5.  LCSC Scraper (mock JSON response)
# ══════════════════════════════════════════════════════════════════════════════

class TestLCSCScraper:

    @pytest.mark.asyncio
    async def test_parse_mock_response(self, monkeypatch):
        from scrapers.lcsc_scraper import LCSCScraper

        mock_data = {
            "result": {
                "productSearchResultVO": {
                    "productList": [
                        {
                            "productModel": "ATMEGA328P-AU",
                            "stockNumber": 5000,
                            "minBuyNumber": 5,
                            "pdfUrl": "https://example.com/datasheet.pdf",
                            "productPriceList": [
                                {"ladder": 5,   "usdPrice": "0.98"},
                                {"ladder": 50,  "usdPrice": "0.85"},
                                {"ladder": 500, "usdPrice": "0.72"},
                            ],
                        }
                    ]
                }
            }
        }

        scraper = LCSCScraper()

        async def fake_get_json(*a, **kw):
            return mock_data

        monkeypatch.setattr(scraper, "_get_json", fake_get_json)
        monkeypatch.setattr(scraper, "_get_cache", lambda *a: None)
        monkeypatch.setattr(scraper, "_set_cache", lambda *a: None)

        results = await scraper.search("ATmega328P", qty=10, rate_inr=83.5)
        assert len(results) == 1
        assert results[0]["vendor"] == "LCSC"
        assert results[0]["unit_price_inr"] == round(0.98 * 83.5, 4)
        assert results[0]["stock_qty"] == 5000
        assert results[0]["moq"] == 5
        assert len(results[0]["price_breaks"]) == 3

    @pytest.mark.asyncio
    async def test_empty_response_returns_empty_list(self, monkeypatch):
        from scrapers.lcsc_scraper import LCSCScraper
        scraper = LCSCScraper()

        async def fake_get_json(*a, **kw):
            return None

        monkeypatch.setattr(scraper, "_get_json", fake_get_json)
        monkeypatch.setattr(scraper, "_get_cache", lambda *a: None)
        monkeypatch.setattr(scraper, "_set_cache", lambda *a: None)

        results = await scraper.search("UNKNOWNPART999", qty=1, rate_inr=83.5)
        assert results == []


# ══════════════════════════════════════════════════════════════════════════════
# 6.  Excel Exporter
# ══════════════════════════════════════════════════════════════════════════════

class TestExcelExporter:

    def _make_results(self):
        return [
            {
                "normalized_mpn": "ATmega328P",
                "component": "ATmega328P-PU",
                "quantity_required": 10,
                "best_vendor": "LCSC",
                "best_unit_price_inr": 81.57,
                "best_total_price_inr": 815.70,
                "availability": "In Stock",
                "moq": 5,
                "savings_vs_worst_inr": 310.50,
                "recommendation_reason": "LCSC: best price, MOQ fits",
                "all_vendors": [
                    {"vendor": "LCSC",    "vendor_part_number": "C1570968", "unit_price_inr": 81.57,  "stock_qty": 12000, "moq": 5,  "availability": "In Stock",     "lead_time_weeks": None, "price_breaks": [{"qty": 5, "price_inr": 81.57}, {"qty": 50, "price_inr": 70.98}]},
                    {"vendor": "DigiKey", "vendor_part_number": "ATMEGA-ND", "unit_price_inr": 104.37, "stock_qty": 5000, "moq": 1,  "availability": "In Stock",     "lead_time_weeks": None, "price_breaks": []},
                    {"vendor": "Arrow",   "vendor_part_number": "ATmega-AR", "unit_price_inr": 112.62, "stock_qty": 0,    "moq": 1,  "availability": "Out of Stock", "lead_time_weeks": 8,    "price_breaks": []},
                ],
            },
        ]

    def test_export_returns_bytes(self):
        from exporter.excel_exporter import export_to_excel
        data = export_to_excel(self._make_results(), "test-task-id", 83.5)
        assert isinstance(data, bytes)
        assert len(data) > 1000   # non-trivial file

    def test_export_is_valid_xlsx(self):
        from exporter.excel_exporter import export_to_excel
        import openpyxl, io
        data = export_to_excel(self._make_results(), "test-task-id", 83.5)
        wb = openpyxl.load_workbook(io.BytesIO(data))
        assert "Optimization Summary" in wb.sheetnames
        assert "Vendor Comparison" in wb.sheetnames
        assert "Audit Log" in wb.sheetnames

    def test_summary_sheet_has_data_row(self):
        from exporter.excel_exporter import export_to_excel
        import openpyxl, io
        data = export_to_excel(self._make_results(), "test-task-id", 83.5)
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["Optimization Summary"]
        # Row 1 = headers, Row 2 = first data row
        assert ws.cell(2, 2).value == "ATmega328P"   # normalized_mpn
        assert ws.cell(2, 3).value == 10              # qty

    def test_vendor_comparison_has_all_vendors(self):
        from exporter.excel_exporter import export_to_excel
        import openpyxl, io
        data = export_to_excel(self._make_results(), "test-task-id", 83.5)
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["Vendor Comparison"]
        vendors_in_sheet = [ws.cell(r, 2).value for r in range(2, ws.max_row + 1) if ws.cell(r, 2).value]
        assert "LCSC" in vendors_in_sheet
        assert "DigiKey" in vendors_in_sheet


# ══════════════════════════════════════════════════════════════════════════════
# 7.  FastAPI endpoints (integration, using TestClient)
# ══════════════════════════════════════════════════════════════════════════════

class TestFastAPIEndpoints:

    @pytest.fixture(autouse=True)
    def setup_test_db(self, tmp_path, monkeypatch):
        """Use a temporary SQLite DB for each test."""
        db_path = str(tmp_path / "test_bom.db")
        monkeypatch.setenv("DATABASE_URL", f"sqlite:///{db_path}")
        monkeypatch.setenv("REDIS_URL", "")  # disable Redis in tests

    def _get_client(self):
        import importlib
        import main as m
        importlib.reload(m)   # pick up env var changes
        from fastapi.testclient import TestClient
        return TestClient(m.app)

    def test_root_returns_ok(self):
        from fastapi.testclient import TestClient
        from main import app
        client = TestClient(app)
        resp = client.get("/")
        assert resp.status_code == 200
        assert resp.json()["currency"] == "INR"

    def test_health_endpoint(self):
        from fastapi.testclient import TestClient
        from main import app
        client = TestClient(app)
        resp = client.get("/health")
        assert resp.status_code == 200

    def test_normalize_endpoint(self):
        from fastapi.testclient import TestClient
        from main import app
        client = TestClient(app)
        resp = client.post("/api/components/normalize", json={"mpn": "ATmega328P-PU"})
        assert resp.status_code == 200
        body = resp.json()
        assert body["normalized"] == "ATmega328P"
        assert body["confidence"] >= 0.90

    def test_normalize_alias(self):
        from fastapi.testclient import TestClient
        from main import app
        client = TestClient(app)
        resp = client.post("/api/components/normalize", json={"mpn": "lm7805"})
        assert resp.status_code == 200
        assert resp.json()["normalized"] == "L7805"

    def test_analyze_returns_task_id(self, monkeypatch):
        from fastapi.testclient import TestClient
        from main import app

        # Patch Celery task to avoid actually running
        monkeypatch.setattr("main.analyze_bom.delay", lambda *a, **kw: None)

        client = TestClient(app)
        resp = client.post("/api/bom/analyze", json={
            "bom_text": "Qty,MPN,Description\n10,ATmega328P-PU,MCU",
            "format": "csv",
        })
        assert resp.status_code == 200
        body = resp.json()
        assert "task_id" in body
        assert len(body["task_id"]) == 36   # UUID4 length

    def test_status_of_nonexistent_task_returns_404(self):
        from fastapi.testclient import TestClient
        from main import app
        client = TestClient(app)
        resp = client.get("/api/bom/status/nonexistent-task-id")
        assert resp.status_code == 404

    def test_analyze_empty_bom_returns_400(self, monkeypatch):
        from fastapi.testclient import TestClient
        from main import app
        monkeypatch.setattr("main.analyze_bom.delay", lambda *a, **kw: None)
        client = TestClient(app)
        resp = client.post("/api/bom/analyze", json={"bom_text": "   ", "format": "csv"})
        assert resp.status_code == 400

    def test_fx_rate_endpoint(self, monkeypatch):
        from fastapi.testclient import TestClient
        from main import app
        monkeypatch.setattr("main.get_usd_inr_rate", lambda: 83.5)
        client = TestClient(app)
        resp = client.get("/api/fx/rate")
        assert resp.status_code == 200
        body = resp.json()
        assert body["rate"] == 83.5
        assert body["to"] == "INR"


# ══════════════════════════════════════════════════════════════════════════════
# Pytest config
# ══════════════════════════════════════════════════════════════════════════════

def pytest_configure(config):
    config.addinivalue_line("markers", "asyncio: mark test as async")
